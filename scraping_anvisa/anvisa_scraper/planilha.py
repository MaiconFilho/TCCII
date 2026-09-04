import re
import unicodedata
from pathlib import Path

from openpyxl import load_workbook

from .modelos import MedicamentoParaColeta


ABA_PADRAO = "Medicamentos monitorados"
COLUNA_NOME = "Nome do medicamento"


def normalizar_nome(valor: object) -> str:
    texto = str(valor or "").strip().casefold()
    texto = unicodedata.normalize("NFKD", texto)
    texto = "".join(caractere for caractere in texto if not unicodedata.combining(caractere))
    return re.sub(r"\s+", " ", texto).strip()


def carregar_medicamentos(
    caminho: Path,
    aba: str = ABA_PADRAO,
) -> list[MedicamentoParaColeta]:
    if not caminho.exists():
        raise FileNotFoundError(f"Planilha não encontrada: {caminho}")

    workbook = load_workbook(caminho, read_only=True, data_only=True)
    try:
        if aba not in workbook.sheetnames:
            raise ValueError(
                f"A aba '{aba}' não existe. Abas: {', '.join(workbook.sheetnames)}"
            )

        sheet = workbook[aba]
        linhas = sheet.iter_rows(values_only=True)
        cabecalho = next(linhas, None)
        if not cabecalho:
            raise ValueError(f"A aba '{aba}' está vazia.")

        indices = {str(nome).strip(): indice for indice, nome in enumerate(cabecalho)}
        if COLUNA_NOME not in indices:
            raise ValueError(f"Coluna obrigatória ausente: {COLUNA_NOME}")

        nomes_em_ordem: list[str] = []
        nomes_originais: dict[str, str] = {}
        ocorrencias: dict[str, int] = {}

        for numero_linha, linha in enumerate(linhas, start=2):
            nome = str(linha[indices[COLUNA_NOME]] or "").strip()
            if not nome:
                if all(valor in (None, "") for valor in linha):
                    continue
                raise ValueError(f"Linha {numero_linha} sem nome do medicamento.")

            nome_normalizado = normalizar_nome(nome)
            if not nome_normalizado:
                continue

            if nome_normalizado not in ocorrencias:
                nomes_em_ordem.append(nome_normalizado)
                nomes_originais[nome_normalizado] = nome
                ocorrencias[nome_normalizado] = 0
            ocorrencias[nome_normalizado] += 1
    finally:
        workbook.close()

    if not nomes_em_ordem:
        raise ValueError(f"Nenhum medicamento válido foi encontrado na aba '{aba}'.")

    return [
        MedicamentoParaColeta(
            nome_produto=nomes_originais[nome_normalizado],
            nome_normalizado=nome_normalizado,
            quantidade_registros_planilha=ocorrencias[nome_normalizado],
        )
        for nome_normalizado in nomes_em_ordem
    ]
